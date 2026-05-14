import os
import re
import io
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURATION: Set Tesseract Path (Uncomment and adjust if on Windows)
# ==============================================================================
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def sanitize_for_excel(text: str) -> str:
    """
    Removes illegal ASCII/Unicode control characters that corrupt Excel XML structures,
    ensuring openpyxl writes data cleanly without throwing IllegalCharacterError.
    """
    if not text or not isinstance(text, str):
        return ""
    
    # Strip raw binary control codes (Removes ASCII 0-31 except tab/newline, and strips 127-159)
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    return cleaned.strip()

def clean_text_payload(text: str) -> str:
    """
    Standardizes spacing and line breaks to prevent merged strings while
    preserving semantic block separation.
    """
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def extract_value_between_anchors(text: str, start_pattern: str, stop_keywords: list) -> str:
    """
    Finds the start pattern, looks forward until it hits one of the stop keywords
    or the end of the string, and returns the cleaned payload in between.
    """
    match = re.search(start_pattern, text, re.IGNORECASE)
    if not match:
        return ""
    
    start_idx = match.end()
    remaining_text = text[start_idx:]
    
    end_idx = len(remaining_text)
    for kw in stop_keywords:
        kw_match = re.search(r'\b' + re.escape(kw) + r'\b', remaining_text, re.IGNORECASE)
        if kw_match and kw_match.start() < end_idx:
            end_idx = kw_match.start()
            
    payload = remaining_text[:end_idx].strip(" :-\n\r\t")
    
    # Reject false-positive captures like raw headers or solitary characters
    if payload.upper() in ["CODE", "NAME", "ITEM", "BRAND", "SIZE", ""] or len(payload) <= 2:
        return ""
        
    return payload

def parse_catalog_text(text: str) -> dict:
    """
    Optimized key-value pointer parsing logic resistant to irregular PDF layouts.
    Updated dynamically to support Decorative Pots documentation workflows.
    """
    # 1. Extract Item Code
    item_code = extract_value_between_anchors(
        text, 
        r'\b(?:ITEM\s*CODE|CODE)\s*[:\-]?\s*', 
        ['BRAND', 'ITEM', 'ITEM NAME', 'SIZE', 'MATERIAL', 'DIMENSIONS', 'FINISH', 'QUANTITY']
    )
    if not item_code:
        # Fallback regex capturing standard SKU formats (e.g., SW25-..., POT-..., numeric codes)
        fallback = re.search(r'\b([A-Z0-9]{2,8}\-[A-Z0-9\-]{3,15}|[0-9]{4,10})\b', text)
        item_code = fallback.group(1) if fallback else "Not Found"

    # 2. Extract Brand Name
    brand_name = extract_value_between_anchors(
        text, 
        r'\bBRAND\s*[:\-]?\s*', 
        ['ITEM', 'ITEM NAME', 'SIZE', 'MATERIAL', 'DIMENSIONS', 'FINISH', 'QUANTITY', 'CODE']
    )
    if not brand_name:
        # Secondary scan for known manufacturers embedded within raw body strings
        upper_text = text.upper()
        for known_brand in ["CASAMIA", "VISIONNAIRE", "LONGHI", "PORRO", "DITRE ITALIA", "VONDOM", "SERRALUNGA", "KHILIA"]:
            if known_brand in upper_text:
                brand_name = known_brand
                break
        if not brand_name:
            brand_name = "Unspecified"

    # 3. Extract Item Name
    item_name = extract_value_between_anchors(
        text, 
        r'\b(?:ITEM\s*NAME|ITEM)\s*[:\-]?\s*', 
        ['SIZE', 'MATERIAL', 'DIMENSIONS', 'FINISH', 'QUANTITY', 'UNIT PRICE', 'STATUS', 'DESCRIPTION', 'CASAMIA']
    )
    
    # Clean up overlaps if the generic "ITEM:" header snagged the brand name by mistake
    if item_name.upper() == brand_name.upper() or not item_name:
        # Contextual search targeting common landscape and planter terminologies
        fallback_name = re.search(r'\b([A-Z\s\|\(\)]+(?:POT|PLANTER|VASE|VASO|BOWL)[A-Z\s\|\(\)]*)\b', text, re.IGNORECASE)
        if fallback_name:
            item_name = fallback_name.group(1).strip()
        else:
            item_name = "Not Found"

    # Return cleanly sanitized payload mapped directly to output keys
    return {
        "Item Code": sanitize_for_excel(item_code),
        "Brand Name": sanitize_for_excel(brand_name),
        "Item Name": sanitize_for_excel(item_name)
    }

def process_pdf_locally(pdf_path: str) -> list[dict]:
    """Iterates through the PDF utilizing hybrid parsing logic."""
    extracted_records = []
    doc = fitz.open(pdf_path)
    print(f"Starting optimized processing of: {pdf_path} ({len(doc)} pages)")

    for page_num in range(len(doc)):
        page = doc[page_num]
        actual_page_num = page_num + 1
        print(f"Processing Page {actual_page_num}...", end=" ")

        raw_text = page.get_text("text")

        # Triage Logic: If native text layer is missing or highly minimal, execute OCR
        if len(raw_text.strip()) < 40 or not any(kw in raw_text.upper() for kw in ["ITEM", "CODE", "SIZE", "POT"]):
            print("[OCR Triggered]", end=" ")
            try:
                pix = page.get_pixmap(dpi=200)
                img_data = Image.open(io.BytesIO(pix.tobytes("png")))
                raw_text = pytesseract.image_to_string(img_data)
            except Exception as e:
                print(f"(Failed: {e})", end=" ")
                raw_text = ""

        cleaned_text = clean_text_payload(raw_text)
        parsed_data = parse_catalog_text(cleaned_text)
        
        extracted_records.append({
            "Page": actual_page_num,
            "Item Code": parsed_data["Item Code"],
            "Brand Name": parsed_data["Brand Name"],
            "Item Name": parsed_data["Item Name"]
        })
        print("Done.")

    return extracted_records

def export_styled_excel(data: list[dict], output_filename: str):
    """Exports structured dictionaries into a polished tech-themed Excel file."""
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Extracted Catalog', index=False, startrow=4)
        
        worksheet = writer.sheets['Extracted Catalog']
        worksheet.views.sheetView[0].showGridLines = True
        
        # Professional Design Aesthetics (Slate Grey & Navy Blue)
        NAVY_BLUE = "1B365D"
        SLATE_GREY = "4A5568"
        LIGHT_SLATE = "F1F5F9"
        BORDER_COLOR = "CBD5E1"
        font_family = "Segoe UI"
        
        thin_border = Border(
            left=Side(style='thin', color=BORDER_COLOR), right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR), bottom=Side(style='thin', color=BORDER_COLOR)
        )
        
        # Title Header
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Decorative Pots Offline Catalog Extraction"
        worksheet["A1"].font = Font(name=font_family, size=16, bold=True, color=NAVY_BLUE)
        worksheet["A2"] = "Optimized block-scanning via PyMuPDF & Tesseract OCR"
        worksheet["A2"].font = Font(name=font_family, size=10, italic=True, color=SLATE_GREY)
        worksheet.row_dimensions[1].height = 35
        worksheet.row_dimensions[2].height = 18
        
        # Table Columns Styling
        worksheet.row_dimensions[5].height = 26
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=5, column=col_idx)
            cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Row Iteration Integration
        for row_idx in range(len(df)):
            current_row = 6 + row_idx
            worksheet.row_dimensions[current_row].height = 20
            row_fill = PatternFill(start_color="FFFFFF" if row_idx % 2 == 0 else LIGHT_SLATE, fill_type="solid")
            
            for col_idx in range(1, 5):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.font = Font(name=font_family, size=10)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
                    
        # Clean Padding Column Autofit
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col if cell.row >= 5)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)

    print(f"\nOptimization Complete! File saved cleanly to: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    import datetime
    
    TARGET_PDF = "OUTDOOR ACCESSORIES.pdf"
    OUTPUT_FILE = "Extracted_OUTDOOR ACCESSORIES_Catalog_Data.xlsx"
    
    if not os.path.exists(TARGET_PDF):
        print(f"Error: Target file '{TARGET_PDF}' not found in the root execution directory.")
    else:
        extracted_data = process_pdf_locally(TARGET_PDF)
        
        # Safe save block ensuring work is preserved if output file is write-locked by OS
        try:
            export_styled_excel(extracted_data, OUTPUT_FILE)
        except PermissionError:
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            safe_filename = f"Extracted_DECORATIVE_POTS_{timestamp}.xlsx"
            print(f"\n[OS Lock Detected] Target spreadsheet is currently open or syncing.")
            print(f"Preserving processed batch directly to alternate output: '{safe_filename}'...")
            export_styled_excel(extracted_data, safe_filename)
import os
import re
import fitz  # PyMuPDF
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def sanitize_for_excel(text: str) -> str:
    """
    Strips raw binary control characters from strings to prevent corrupting 
    the XML schema of the final Excel file, ensuring openpyxl stability.
    """
    if not text or not isinstance(text, str):
        return ""
    # Strip raw binary control codes (Removes ASCII 0-31 except tab/newline, and strips 127-159)
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    return cleaned.strip()

def process_showroom_pdf_complete(pdf_path: str) -> list[dict]:
    """
    Executes a stateful matrix scan natively via PyMuPDF. Evaluates both primary 
    tracking columns simultaneously to intercept line-break arrays and floating-point 
    SKU shifts without dropping inventory records.
    """
    records = []
    doc = fitz.open(pdf_path)
    print(f"Starting localized multi-column extraction pipeline for: '{pdf_path}' ({len(doc)} pages)")
    
    # Establish persistent contextual state tracking
    current_doc_no = "KA-1"
    
    # Robust matrix exclusion array preventing descriptive labels from mapping as SKUs
    ignore_words = {
        "doc", "model", "no", "photo", "description", "unit", "qty", "price", "cust", "om", "aed",
        "pcs", "set", "sets", "damage", "damaged", "available", "product", "name", "main", "colour",
        "special", "requirements", "client", "meter", "sqm", "roller", "kg", "pot", "vase", "can",
        "box", "tray", "clock", "shelf", "light", "lamp", "chair", "table", "painting", "decoration",
        "bottle", "book", "cup", "leaf", "branch", "mirrow", "mirror", "cushion", "stool", "timer",
        "plate", "compass", "telescope", "handgun", "shaker", "bulb", "indicator", "storage", "bag",
        "rack", "cabinet", "calender", "calendar", "board", "arrows", "word", "panel", "disk",
        "drawer", "map", "tree", "lantern", "game", "opener", "vessel", "peace", "piece", "case",
        "seed", "mesh", "ring", "earthbags", "trolley", "sofa", "chest", "desk", "ottoman", "block",
        "brick", "receptacle", "ball", "dandelion", "holder", "picture", "material", "size",
        "selling", "total", "showroom", "unit selling", "total selling", "ashi", "loft", "stop"
    }
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        actual_page = page_num + 1
        print(f"--> Scanning Page {actual_page}...", end=" ")
        
        # Native document table structure retrieval
        tabs = page.find_tables()
        if not tabs.tables:
            print("Skipped (Zero layout matrices found).")
            continue
            
        page_records_count = 0
        
        for table in tabs:
            df = table.to_pandas()
            
            for _, row in df.iterrows():
                # Format cell buffers safely as standard strings
                cells = [str(c) if pd.notna(c) else "" for c in row]
                if not cells or not cells[0].strip():
                    continue
                    
                # Layout triage check: Skip literal table matrix definition headers
                c0_clean = re.sub(r'\s+', ' ', cells[0]).strip().lower()
                if any(c0_clean.startswith(h) for h in ["cust", "doc no", "item#", "model no"]):
                    continue
                    
                # Step 1: Intercept persistent DOC NO layout markers natively across Col 0 and Col 1
                for col_val in cells[:2]:
                    if not col_val.strip():
                        continue
                    doc_matches = list(re.finditer(r'\b(KA[\-\.]\d+(?:\.\d+)?)\b', col_val, re.IGNORECASE))
                    if doc_matches:
                        # Standardize internal layout formatting cleanly
                        current_doc_no = doc_matches[-1].group(1).upper().replace(".", "-")
                        
                # Step 2: Extract target SKU tokens across primary alignment columns
                extracted_models = []
                
                # Check Column 0 Array Blocks
                val0 = cells[0].strip()
                if val0 and val0.upper() not in ["KA", "KA-"] and not re.fullmatch(r'KA[\-\.]\d+(?:\.\d+)?', val0, re.IGNORECASE):
                    for sub_line in val0.split("\n"):
                        sub_clean = sub_line.strip()
                        if not sub_clean:
                            continue
                        # Decouple context tags if they share the specific line string
                        sub_stripped = re.sub(r'\bKA[\-\.]\d+(?:\.\d+)?\b', '', sub_clean, flags=re.IGNORECASE).strip()
                        if sub_stripped and len(sub_stripped) > 2 and not any(w == sub_stripped.lower() for w in ignore_words):
                            # Validate string contains core alphanumeric inventory keys
                            if any(ch.isdigit() for ch in sub_stripped) or re.match(r'^[A-Z]{2,}\-', sub_stripped):
                                if sub_stripped not in extracted_models:
                                    extracted_models.append(sub_stripped)
                                    
                # Check Column 1 Array Blocks (Catches shifted floats like 5231790.0)
                if len(cells) > 1:
                    val1 = cells[1].strip()
                    if val1:
                        for sub_line in val1.split("\n"):
                            sub_clean = sub_line.strip()
                            # Convert numerical SKU floats back to clean integer strings safely
                            if sub_clean.endswith(".0"):
                                sub_clean = sub_clean[:-2]
                            if sub_clean and len(sub_clean) > 2 and not any(w == sub_clean.lower() for w in ignore_words):
                                if any(ch.isdigit() for ch in sub_clean) or re.match(r'^[A-Z]{2,}\-', sub_clean):
                                    if sub_clean not in extracted_models:
                                        extracted_models.append(sub_clean)
                                        
                # Map structured extractions directly to current page and operational document bounds
                for model_no in extracted_models:
                    records.append({
                        "Page": actual_page,
                        "DOC NO": current_doc_no,
                        "Model Number": sanitize_for_excel(model_no)
                    })
                    page_records_count += 1
                    
        print(f"Mapped {page_records_count} SKUs.")
        
    # Drop edge-case duplication loops directly in layout dataframe structures
    final_records = pd.DataFrame(records).drop_duplicates(subset=["Model Number"]).to_dict(orient="records")
    return final_records

def export_styled_excel(data: list[dict], output_filename: str):
    """Compiles extracted record mappings into an optimized, tech-themed spreadsheet."""
    if not data:
        print("\n[Warning] Script processing returned zero record extractions. Halting compilation phase.")
        return
        
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Extracted Catalog', index=False, startrow=4)
        
        worksheet = writer.sheets['Extracted Catalog']
        worksheet.views.sheetView[0].showGridLines = True
        
        # Professional Corporate Palette (Navy Blue & Slate Grey)
        NAVY_BLUE = "1B365D"
        SLATE_GREY = "4A5568"
        LIGHT_SLATE = "F1F5F9"
        BORDER_COLOR = "CBD5E1"
        font_family = "Segoe UI"
        
        thin_border = Border(
            left=Side(style='thin', color=BORDER_COLOR), right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR), bottom=Side(style='thin', color=BORDER_COLOR)
        )
        
        # Title Header Formatting Block
        worksheet.merge_cells("A1:C1")
        worksheet["A1"] = "Showroom Inventory SKU Extraction"
        worksheet["A1"].font = Font(name=font_family, size=16, bold=True, color=NAVY_BLUE)
        worksheet["A2"] = f"Total Parsed Model Inventory: {len(df)} Unique Items"
        worksheet["A2"].font = Font(name=font_family, size=10, italic=True, color=SLATE_GREY)
        worksheet.row_dimensions[1].height = 35
        worksheet.row_dimensions[2].height = 18
        
        # Style Header Columns
        worksheet.row_dimensions[5].height = 26
        for col_idx in range(1, 4):
            cell = worksheet.cell(row=5, column=col_idx)
            cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Data Loop Alignment Integration
        for row_idx in range(len(df)):
            current_row = 6 + row_idx
            worksheet.row_dimensions[current_row].height = 20
            row_fill = PatternFill(start_color="FFFFFF" if row_idx % 2 == 0 else LIGHT_SLATE, fill_type="solid")
            
            for col_idx in range(1, 4):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.font = Font(name=font_family, size=10)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2] else "left", vertical="center")
                    
        # Calculate Padding Column Dimensions
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col if cell.row >= 5)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)

    print(f"\nOptimization Execution Complete! Output artifact saved securely to: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    import datetime
    
    # Target alignment configuration paths setting
    TARGET_PDF = "ACCESSORIES SHOWROOM-1-45.pdf"  
    OUTPUT_FILE = "Extracted_Showroom_Complete_Data.xlsx"
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    full_pdf_path = os.path.join(current_dir, TARGET_PDF)
    full_out_path = os.path.join(current_dir, OUTPUT_FILE)
    
    if not os.path.exists(full_pdf_path):
        print(f"\n[Execution Failure] Source document unavailable:")
        print(f"👉 Target File Path: '{full_pdf_path}'")
        print("Verify filename exactness and ensure target documentation resides inside root execution root.")
    else:
        extracted_data = process_showroom_pdf_complete(full_pdf_path)
        
        # Exception handler preventing permissions crashing during directory cloud synchronization
        try:
            export_styled_excel(extracted_data, full_out_path)
        except PermissionError:
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            safe_filename = os.path.join(current_dir, f"Extracted_Showroom_Batch_{timestamp}.xlsx")
            print(f"\n[OS Write Lock] Primary write access blocked by operating system threads.")
            print(f"Bypassing active lock directly to secure target duplicate: '{safe_filename}'...")
            export_styled_excel(extracted_data, safe_filename)
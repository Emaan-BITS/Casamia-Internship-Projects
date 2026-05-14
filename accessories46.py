import os
import re
import io
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURATION: Set Tesseract Path (Uncomment and adjust if on Windows)
# ==============================================================================
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def sanitize_for_excel(text: str) -> str:
    """
    Strips raw binary/control characters from strings to prevent corrupting 
    the XML schema of the final Excel file, ensuring openpyxl stability.
    """
    if not text or not isinstance(text, str):
        return ""
    # Strip raw binary control codes (Removes ASCII 0-31 except tab/newline, and strips 127-159)
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    return cleaned.strip()

def clean_text_payload(text: str) -> str:
    """Standardizes spacing and line breaks for baseline parsing fallback."""
    return re.sub(r'\s+', ' ', text).strip()

def extract_value_between_anchors(text: str, start_pattern: str, stop_keywords: list) -> str:
    """Fallback sliding window search for unstructured catalog pages."""
    match = re.search(start_pattern, text, re.IGNORECASE)
    if not match:
        return ""
    
    start_idx = match.end()
    remaining_text = text[start_idx:]
    
    end_idx = len(remaining_text)
    for kw in stop_keywords:
        kw_match = re.search(r'\b' + re.escape(kw) + r'\b', remaining_text, re.IGNORECASE)
        if kw_match and kw_match.start() < end_idx:
            end_idx = kw_match.start()
            
    payload = remaining_text[:end_idx].strip(" :-\n\r\t")
    if payload.upper() in ["CODE", "NAME", "ITEM", "BRAND", "SIZE", "QTY", "MATERIAL", ""] or len(payload) <= 1:
        return ""
    return payload

def parse_tabular_matrix(table_df: pd.DataFrame, actual_page_num: int) -> list[dict]:
    """
    Processes multi-page PDF matrix blocks natively. Uses semantic string 
    scanning to skip layout headers without losing valid starting SKU rows.
    """
    records = []
    
    # Universal array keywords defining layout boundaries across files
    header_keywords = ['item#', 'itemcode', 'item code', 'desc.', 'picture', 'material', 'brand', 'qty', 'description']
    
    for row_idx, row in table_df.iterrows():
        # Clean up cell values uniformly
        cells = [sanitize_for_excel(str(c)) if pd.notna(c) else "" for c in row]
        
        # Ensure row contains viable tracking data
        if len(cells) < 2:
            continue
            
        first_cell_str = str(cells[0]).strip().lower()
        
        # Semantic Boundary Guard: Skip if the cell defines a header label
        if any(first_cell_str.startswith(kw) or first_cell_str == kw for kw in header_keywords):
            continue
            
        # Extract base code string cleanly
        raw_code = cells[0].split("\n")[0].strip()
        
        # Skip empty visual/structural padding rows
        if not raw_code or raw_code.upper() in ["DAMAGED", "SOLD", "NO PHOTO"]:
            continue
            
        # Dynamic Column Assignment based on table layout matrix mapping
        item_name = "Not Found"
        if len(cells) > 1 and cells[1].strip():
            item_name = cells[1].replace("\n", " ").strip()
        elif len(cells) > 2 and cells[2].strip():
            item_name = cells[2].replace("\n", " ").strip()
            
        # If code column contains unparsed arrays, run targeted regex extraction
        if len(raw_code) < 3 or " " in raw_code:
            sku_match = re.search(r'\b([A-Z0-9]{2,12}\-?[A-Z0-9]{2,10})\b', raw_code)
            if sku_match:
                raw_code = sku_match.group(1)
                
        # Consistently filter target item names
        if item_name.upper() == raw_code.upper() or len(item_name) <= 2:
            item_name = "Accessories Item"
            
        records.append({
            "Page": actual_page_num,
            "Item Code": raw_code,
            "Brand Name": "Casamia Showroom", # Catalog collection source context mapping
            "Item Name": item_name
        })
        
    return records

def parse_freeform_text(text: str) -> dict:
    """Fallback structural parser handling unmapped freeform text pages."""
    item_code = extract_value_between_anchors(
        text, r'\b(?:ITEM\s*#|ITEM\s*CODE|CODE)\s*[:\-]?\s*', 
        ['DESC', 'PICTURE', 'MATERIAL', 'SIZE', 'QTY', 'UNIT SELLING', 'TOTAL', 'BRAND']
    )
    if not item_code:
        fallback = re.search(r'\b([A-Z0-9]{3,10}\-?[A-Z0-9]{2,8})\b', text)
        item_code = fallback.group(1) if fallback else "Not Found"

    item_name = extract_value_between_anchors(
        text, r'\b(?:DESC\.|DESCRIPTION|ITEM\s*NAME|ITEM)\s*[:\-]?\s*', 
        ['PICTURE', 'MATERIAL', 'SIZE', 'QTY', 'UNIT SELLING', 'TOTAL', 'AED', 'CASAMIA']
    )
    if not item_name or item_name.upper() == item_code.upper():
        fallback_name = re.search(r'\b([A-Z\s]+(?:POT|VASE|LIGHT|GLASS|PLATE|CHAIR|DESK|TRAY|BOX|ORNAMENT)[A-Z\s]*)\b', text, re.IGNORECASE)
        item_name = fallback_name.group(1).strip() if fallback_name else "Accessories Item"

    return {
        "Item Code": sanitize_for_excel(item_code),
        "Brand Name": "Casamia Showroom",
        "Item Name": sanitize_for_excel(item_name)
    }

def process_pdf_locally(pdf_path: str) -> list[dict]:
    """Processes document payload leveraging unified table/text offline engines."""
    extracted_records = []
    doc = fitz.open(pdf_path)
    print(f"Starting optimized multi-page Table extraction for: {pdf_path} ({len(doc)} pages)")

    for page_num in range(len(doc)):
        page = doc[page_num]
        actual_page_num = page_num + 1
        print(f"Processing Page {actual_page_num}...", end=" ")

        # 1. Native Matrix Parsing Phase
        tabs = page.find_tables()
        if tabs.tables:
            print("[Table Detected] Parsing cells...", end=" ")
            for table in tabs:
                df = table.to_pandas()
                parsed_rows = parse_tabular_matrix(df, actual_page_num)
                extracted_records.extend(parsed_rows)
            print(f"Done ({len(parsed_rows)} SKUs mapped).")
            continue

        # 2. Textual Fallback Interception Phase
        raw_text = page.get_text("text")
        if len(raw_text.strip()) < 40 or not any(kw in raw_text.upper() for kw in ["ITEM", "CODE", "QTY", "AED"]):
            print("[OCR Triggered]", end=" ")
            try:
                pix = page.get_pixmap(dpi=200)
                img_data = Image.open(io.BytesIO(pix.tobytes("png")))
                raw_text = pytesseract.image_to_string(img_data)
            except Exception as e:
                print(f"(Failed: {e})", end=" ")
                raw_text = ""

        cleaned_text = clean_text_payload(raw_text)
        parsed_data = parse_freeform_text(cleaned_text)
        
        extracted_records.append({
            "Page": actual_page_num,
            "Item Code": parsed_data["Item Code"],
            "Brand Name": parsed_data["Brand Name"],
            "Item Name": parsed_data["Item Name"]
        })
        print("Done (Freeform parsed).")

    return extracted_records

def export_styled_excel(data: list[dict], output_filename: str):
    """Compiles extracted arrays into an optimized, cleanly formatted Excel layout."""
    if not data:
        print("\n[Warning] Batch processing returned zero records. Halting export flow.")
        return
        
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Extracted Catalog', index=False, startrow=4)
        
        worksheet = writer.sheets['Extracted Catalog']
        worksheet.views.sheetView[0].showGridLines = True
        
        # Professional UI Palette (Navy Blue & Slate Grey)
        NAVY_BLUE = "1B365D"
        SLATE_GREY = "4A5568"
        LIGHT_SLATE = "F1F5F9"
        BORDER_COLOR = "CBD5E1"
        font_family = "Segoe UI"
        
        thin_border = Border(
            left=Side(style='thin', color=BORDER_COLOR), right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR), bottom=Side(style='thin', color=BORDER_COLOR)
        )
        
        # Build Title Block
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Showroom Accessories Multi-Page Extraction"
        worksheet["A1"].font = Font(name=font_family, size=16, bold=True, color=NAVY_BLUE)
        worksheet["A2"] = "Dynamic matrix extraction mapped locally via PyMuPDF Tables"
        worksheet["A2"].font = Font(name=font_family, size=10, italic=True, color=SLATE_GREY)
        worksheet.row_dimensions[1].height = 35
        worksheet.row_dimensions[2].height = 18
        
        # Style Header Rows
        worksheet.row_dimensions[5].height = 26
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=5, column=col_idx)
            cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Data Loop Application
        for row_idx in range(len(df)):
            current_row = 6 + row_idx
            worksheet.row_dimensions[current_row].height = 20
            row_fill = PatternFill(start_color="FFFFFF" if row_idx % 2 == 0 else LIGHT_SLATE, fill_type="solid")
            
            for col_idx in range(1, 5):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.font = Font(name=font_family, size=10)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
                    
        # Column Autofit Execution
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col if cell.row >= 5)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)

    print(f"\nOptimization Execution Complete! File output successfully generated: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    import datetime
    
    # Active execution parameters targeting specific documentation directory context
    TARGET_PDF = "ACCESSORIES SHOWROOM-46-82.pdf"  
    OUTPUT_FILE = "Extracted_Showroom_Accessories_Data.xlsx"
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    full_pdf_path = os.path.join(current_dir, TARGET_PDF)
    full_out_path = os.path.join(current_dir, OUTPUT_FILE)
    
    if not os.path.exists(full_pdf_path):
        print(f"\n[Error] Unable to locate target document payload:")
        print(f"👉 Expected Path: '{full_pdf_path}'")
        print("Verify filename exactness and ensure active run path matches project execution root.")
    else:
        extracted_data = process_pdf_locally(full_pdf_path)
        
        # Exception handler preventing permissions crashing during local directory synchronization
        try:
            export_styled_excel(extracted_data, full_out_path)
        except PermissionError:
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            safe_filename = os.path.join(current_dir, f"Extracted_Showroom_Batch_{timestamp}.xlsx")
            print(f"\n[OS Write Lock] Target destination spreadsheet is actively engaged by external threads.")
            print(f"Routing compiled buffer directly to alternative safe save target: '{safe_filename}'...")
            export_styled_excel(extracted_data, safe_filename)
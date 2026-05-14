import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# Excel location
input_folder = r"C:\Users\m.emaan\OneDrive - Casa Mia LLC\Automation Task\input"

# Images location
IMAGE_FOLDER = r"C:\Users\m.emaan\OneDrive - Casa Mia LLC\Automation Task\OneDrive_2026-04-20\Image Dump"
PLACEHOLDER_IMAGE_LOCATION = r"C:\Users\m.emaan\OneDrive - Casa Mia LLC\Automation Task\placeholder_image\Screenshot 2026-04-20 111901.png"

for filename in os.listdir(input_folder):
    wb = load_workbook(f"{input_folder}/{filename}")
    sheets = ['Sheet1']
    #           'Carpet', 'Other Furniture']
    #sheets = ['Slabs (40)', 'Washbasin (40)', 'Plumbing (40)', 'Shower Area (40)', 'Toilet (40)', 'Sanitary Accessories (40)']
    #sheets = ['To Order']
    #sheet = wb["Sheet1"]

    #for sheet in wb.worksheets: # use this to make changes to every sheet in wb
    for page in sheets:  # add sheets to be edited in sheets var
    #     if page == "Slabs (40)":
    #         IMAGE_COLUMN_CHARACTER = 'U'
    #     # elif page == "Sanitary Accessories":
    #     #     IMAGE_COLUMN_CHARACTER = 'J'
    #     else:
    #         IMAGE_COLUMN_CHARACTER = 'T'
        IMAGE_COLUMN_CHARACTER = 'E'

        sheet = wb[page]

        for row in range(2, sheet.max_row + 1):
        # for row in range(2, 5000):
            sheet.row_dimensions[row].height = 100
            sheet.column_dimensions[IMAGE_COLUMN_CHARACTER].width = 50

            item_id_cell = sheet[f"A{row}"]
            if item_id_cell.value:
                item_id = item_id_cell.value

                try:
                    # pilimage = PILImage.open(fr"{IMAGE_FOLDER}\{item_id}.jpeg")
                    # width, height = pilimage.size
                    # new_size = (width // 4, height // 4)
                    # img = pilimage.resize(new_size)
                    img = XLImage(fr"{IMAGE_FOLDER}\#{item_id}.jpeg")
                    print(item_id)
                    # img = XLImage(img)
                except FileNotFoundError:
                    img = XLImage(PLACEHOLDER_IMAGE_LOCATION)
                    print(item_id)
                except OSError:
                    print('error')
                    img = XLImage(PLACEHOLDER_IMAGE_LOCATION)
                    print(item_id)
                # as_ratio = img.width / img.height
                img.width = 80
                img.height = 60

                sheet.add_image(img, f"{IMAGE_COLUMN_CHARACTER}{row}")

        print(f"{sheet} done")

    wb.save(f"Out {filename}")


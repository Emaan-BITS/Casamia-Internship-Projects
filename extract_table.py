import os
import re
import io
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURATION: Set Tesseract Path (Uncomment and adjust if on Windows)
# ==============================================================================
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def sanitize_for_excel(text: str) -> str:
    """
    Removes illegal ASCII/Unicode control characters that corrupt Excel XML structures,
    ensuring openpyxl writes data cleanly without throwing IllegalCharacterError.
    """
    if not text or not isinstance(text, str):
        return ""
    # Strip raw binary control codes (Removes ASCII 0-31 except tab/newline, and strips 127-159)
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    return cleaned.strip()

def clean_text_payload(text: str) -> str:
    """Standardizes spacing and line breaks for linear parsing fallback."""
    return re.sub(r'\s+', ' ', text).strip()

def extract_value_between_anchors(text: str, start_pattern: str, stop_keywords: list) -> str:
    """Fallback sliding window logic for pages missing explicit tabular structures."""
    match = re.search(start_pattern, text, re.IGNORECASE)
    if not match:
        return ""
    
    start_idx = match.end()
    remaining_text = text[start_idx:]
    
    end_idx = len(remaining_text)
    for kw in stop_keywords:
        kw_match = re.search(r'\b' + re.escape(kw) + r'\b', remaining_text, re.IGNORECASE)
        if kw_match and kw_match.start() < end_idx:
            end_idx = kw_match.start()
            
    payload = remaining_text[:end_idx].strip(" :-\n\r\t")
    if payload.upper() in ["CODE", "NAME", "ITEM", "BRAND", "SIZE", "QTY", ""] or len(payload) <= 2:
        return ""
    return payload

def parse_tabular_matrix(table_df: pd.DataFrame, actual_page_num: int) -> list[dict]:
    """
    Directly extracts records from native document tables.
    Maps columns dynamically based on standard matrix formatting.
    """
    records = []
    # Drop rows that are clearly header definitions
    header_keywords = ['ItemCode', 'Item Image', 'Item Desc', 'Brand', 'Collection']
    
    for _, row in table_df.iterrows():
        # Clean array values
        row_cells = [sanitize_for_excel(str(cell)) if pd.notna(cell) else "" for cell in row]
        
        # Verify row contains substantial string length and isn't a top-level header row
        if len(row_cells) < 3 or any(kw.lower() in str(row_cells[0]).lower() for kw in header_keywords):
            continue
            
        item_code = row_cells[0].split("\n")[0].strip()
        item_desc = row_cells[2].strip() if len(row_cells) > 2 else ""
        brand_col = row_cells[3].strip() if len(row_cells) > 3 else ""
        
        # Skip empty layout spacer rows
        if not item_code and not item_desc:
            continue
            
        # Standardize Item Code extraction
        if not item_code or len(item_code) < 3:
            code_match = re.search(r'\b([A-Z0-9]{1,4}\.[A-Z0-9]{3}\.[A-Z0-9]{3}\.[A-Z0-9]{3})\b', item_desc)
            item_code = code_match.group(1) if code_match else "Not Found"
            
        # Parse clear item details from description strings
        item_name = item_desc.split(",")[0].strip() if item_desc else "Not Found"
        
        # Identify manufacturer brands cleanly
        brand_name = brand_col.split("\n")[0].strip() if brand_col else ""
        if not brand_name or brand_name in ["Bed Sheet/Cover/Linen", "Cushion", "Frame", "N/A"]:
            if "Gianfranco Ferre" in item_desc:
                brand_name = "Gianfranco Ferre"
            else:
                brand_name = "Unspecified"
                
        records.append({
            "Page": actual_page_num,
            "Item Code": item_code,
            "Brand Name": brand_name,
            "Item Name": item_name
        })
        
    return records

def parse_freeform_text(text: str) -> dict:
    """Fallback block parsing for non-tabular layouts."""
    item_code = extract_value_between_anchors(
        text, r'\b(?:ITEM\s*CODE|CODE)\s*[:\-]?\s*', 
        ['BRAND', 'ITEM', 'ITEM NAME', 'SIZE', 'MATERIAL', 'DIMENSIONS', 'FINISH', 'QTY', 'QUANTITY']
    )
    if not item_code:
        fallback = re.search(r'\b([A-Z0-9]{1,4}\.[A-Z0-9]{3}\.[A-Z0-9]{3}\.[A-Z0-9]{3}|[A-Z0-9]{2,8}\-[A-Z0-9\-]{3,15})\b', text)
        item_code = fallback.group(1) if fallback else "Not Found"

    brand_name = extract_value_between_anchors(
        text, r'\bBRAND\s*[:\-]?\s*', 
        ['ITEM', 'ITEM NAME', 'SIZE', 'MATERIAL', 'DIMENSIONS', 'FINISH', 'QTY', 'CODE']
    )
    if not brand_name:
        for known_brand in ["Gianfranco Ferre", "GF FERRE", "CASAMIA", "VISIONNAIRE", "LONGHI", "PORRO"]:
            if known_brand.upper() in text.upper():
                brand_name = known_brand
                break
        if not brand_name: brand_name = "Unspecified"

    item_name = extract_value_between_anchors(
        text, r'\b(?:ITEM\s*NAME|ITEM|DESC)\s*[:\-]?\s*', 
        ['SIZE', 'MATERIAL', 'DIMENSIONS', 'FINISH', 'QTY', 'QUANTITY', 'UNIT PRICE', 'STATUS', 'COLOUR', 'AED']
    )
    if item_name.upper() == brand_name.upper() or not item_name:
        fallback_name = re.search(r'\b([A-Z\s\|\(\)]+(?:QUILT|CUSHION|PILLOW|BED|TOWEL|VASE|GLASS|PLATE|FRAME)[A-Z\s\|\(\)]*)\b', text, re.IGNORECASE)
        item_name = fallback_name.group(1).strip() if fallback_name else "Not Found"

    return {
        "Item Code": sanitize_for_excel(item_code),
        "Brand Name": sanitize_for_excel(brand_name),
        "Item Name": sanitize_for_excel(item_name)
    }

def process_pdf_locally(pdf_path: str) -> list[dict]:
    """
    Hybrid Execution Pipeline: Prioritizes explicit PDF Table Matrixes.
    Falls back smoothly to standard text logic if grid matrixes are absent.
    """
    extracted_records = []
    doc = fitz.open(pdf_path)
    print(f"Starting Tabular Matrix extraction pipeline for: {pdf_path} ({len(doc)} pages)")

    for page_num in range(len(doc)):
        page = doc[page_num]
        actual_page_num = page_num + 1
        print(f"Processing Page {actual_page_num}...", end=" ")

        # 1. Check for native tabular document formatting
        tabs = page.find_tables()
        if tabs.tables:
            print("[Native Table Detected] Extracting Matrix...", end=" ")
            for table in tabs:
                df = table.to_pandas()
                parsed_table_records = parse_tabular_matrix(df, actual_page_num)
                extracted_records.extend(parsed_table_records)
            print(f"Done ({len(parsed_table_records)} rows mapped).")
            continue

        # 2. Fall back to freeform parsing logic if grid tables are missing
        raw_text = page.get_text("text")
        if len(raw_text.strip()) < 40 or not any(kw in raw_text.upper() for kw in ["ITEM", "CODE", "QUILT", "CUSHION", "PLATE"]):
            print("[OCR Triggered]", end=" ")
            try:
                pix = page.get_pixmap(dpi=200)
                img_data = Image.open(io.BytesIO(pix.tobytes("png")))
                raw_text = pytesseract.image_to_string(img_data)
            except Exception as e:
                print(f"(Failed: {e})", end=" ")
                raw_text = ""

        cleaned_text = clean_text_payload(raw_text)
        parsed_data = parse_freeform_text(cleaned_text)
        
        extracted_records.append({
            "Page": actual_page_num,
            "Item Code": parsed_data["Item Code"],
            "Brand Name": parsed_data["Brand Name"],
            "Item Name": parsed_data["Item Name"]
        })
        print("Done (Freeform block parsed).")

    return extracted_records

def export_styled_excel(data: list[dict], output_filename: str):
    """Exports structured data dictionary into a tech-branded spreadsheet."""
    if not data:
        print("\n[Warning] No records extracted. Terminating spreadsheet generation phase.")
        return
        
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Extracted Catalog', index=False, startrow=4)
        
        worksheet = writer.sheets['Extracted Catalog']
        worksheet.views.sheetView[0].showGridLines = True
        
        # Professional UI Palette (Navy Blue & Slate Grey)
        NAVY_BLUE = "1B365D"
        SLATE_GREY = "4A5568"
        LIGHT_SLATE = "F1F5F9"
        BORDER_COLOR = "CBD5E1"
        font_family = "Segoe UI"
        
        thin_border = Border(
            left=Side(style='thin', color=BORDER_COLOR), right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR), bottom=Side(style='thin', color=BORDER_COLOR)
        )
        
        # Title Block Creation
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Home Accessories Offline Matrix Extraction"
        worksheet["A1"].font = Font(name=font_family, size=16, bold=True, color=NAVY_BLUE)
        worksheet["A2"] = "Multi-engine tabular parsing via PyMuPDF Tables & Tesseract OCR"
        worksheet["A2"].font = Font(name=font_family, size=10, italic=True, color=SLATE_GREY)
        worksheet.row_dimensions[1].height = 35
        worksheet.row_dimensions[2].height = 18
        
        # Header Row Injection
        worksheet.row_dimensions[5].height = 26
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=5, column=col_idx)
            cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Data Cell Loop Alignment
        for row_idx in range(len(df)):
            current_row = 6 + row_idx
            worksheet.row_dimensions[current_row].height = 20
            row_fill = PatternFill(start_color="FFFFFF" if row_idx % 2 == 0 else LIGHT_SLATE, fill_type="solid")
            
            for col_idx in range(1, 5):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.font = Font(name=font_family, size=10)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
                    
        # Precise Padding Column Sizing
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col if cell.row >= 5)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)

    print(f"\nMatrix Optimization Complete! Output generated cleanly: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    import datetime
    
    # Target alignment configuration setting
    TARGET_PDF = "GF FERRE HOME-ACCESSORIES.pdf"  
    OUTPUT_FILE = "Extracted_GF_FERRE_Accessories_Data.xlsx"
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    full_pdf_path = os.path.join(current_dir, TARGET_PDF)
    full_out_path = os.path.join(current_dir, OUTPUT_FILE)
    
    if not os.path.exists(full_pdf_path):
        print(f"\n[Execution Failure] Source payload unavailable:")
        print(f"👉 Checked Directory Path: '{full_pdf_path}'")
        print("Ensure target documentation matches exactly and resides inside root script execution context.")
    else:
        extracted_data = process_pdf_locally(full_pdf_path)
        
        # OS Lock interception processing flow
        try:
            export_styled_excel(extracted_data, full_out_path)
        except PermissionError:
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            safe_filename = os.path.join(current_dir, f"Extracted_GF_FERRE_{timestamp}.xlsx")
            print(f"\n[OS Lock Detected] Primary write access blocked by operating system threads.")
            print(f"Bypassing active file block directly to secure mirror backup: '{safe_filename}'...")
            export_styled_excel(extracted_data, safe_filename)